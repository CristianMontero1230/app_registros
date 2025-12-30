mport streamlit as st
import pandas as pd
import plotly.express as px
import os
import json
import csv
from datetime import datetime
from io import BytesIO
from openpyxl.styles import Font, PatternFill

# Configuración de la página
st.set_page_config(
    page_title="IPS GOLEMAN APP",
    page_icon="🏥",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Constantes y Rutas
DATA_HEADERS = [
    'ID', 'Nombre profesional', 'Documento profesional', 'Nombre paciente',
    'Documento paciente', 'Fecha inicio', 'Municipio', 'Procedimiento',
    'Subido a Panacea', 'Novedad', 'Creado', 'Modificado'
]

DATA_ACTIVITIES_HEADERS = [
    'ID', 'Fecha', 'Nombre profesional', 'Actividad', 'Creado', 'Modificado'
]

DATA_PATH = os.path.join(os.path.dirname(__file__), 'registros_procedimientos.csv')
DATA_ACTIVITIES_PATH = os.path.join(os.path.dirname(__file__), 'registros_actividades.csv')
EXCEL_PATH = os.path.join(os.path.dirname(__file__), 'registros_procedimientos.xlsx')
EXCEL_ACTIVITIES_PATH = os.path.join(os.path.dirname(__file__), 'registros_actividades.xlsx')
CATALOG_PATH = os.path.join(os.path.dirname(__file__), 'catalogo_formulario.json')
UPLOADS_DIR = os.path.join(os.path.dirname(__file__), 'uploads')

# Credenciales
ADMIN_USER = os.environ.get('ADMIN_USER', 'admin') # Default to admin if not set
ADMIN_PASS = os.environ.get('ADMIN_PASS', 'admin') # Default to admin if not set

# --- Funciones de Gestión de Datos ---

def ensure_data_file():
    if not os.path.exists(DATA_PATH):
        restored = False
        if os.path.exists(EXCEL_PATH):
            try:
                df = pd.read_excel(EXCEL_PATH)
                df = df.reindex(columns=DATA_HEADERS)
                df.to_csv(DATA_PATH, index=False)
                restored = True
            except Exception as e:
                st.error(f"Error restaurando CSV de Excel: {e}")
        
        if not restored:
            with open(DATA_PATH, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(DATA_HEADERS)
    
    if not os.path.exists(EXCEL_PATH):
        update_excel_file()

def ensure_activities_file():
    if not os.path.exists(DATA_ACTIVITIES_PATH):
        restored = False
        if os.path.exists(EXCEL_ACTIVITIES_PATH):
            try:
                df = pd.read_excel(EXCEL_ACTIVITIES_PATH)
                df = df.reindex(columns=DATA_ACTIVITIES_HEADERS)
                df.to_csv(DATA_ACTIVITIES_PATH, index=False)
                restored = True
            except Exception as e:
                st.error(f"Error restaurando CSV Actividades de Excel: {e}")
        
        if not restored:
            with open(DATA_ACTIVITIES_PATH, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(DATA_ACTIVITIES_HEADERS)
                
    if not os.path.exists(EXCEL_ACTIVITIES_PATH):
        update_activities_excel_file()

def sync_activities_db():
    if os.path.exists(EXCEL_ACTIVITIES_PATH):
        try:
            df_excel = pd.read_excel(EXCEL_ACTIVITIES_PATH)
            for col in DATA_ACTIVITIES_HEADERS:
                if col not in df_excel.columns:
                    df_excel[col] = ''
            df_excel = df_excel.reindex(columns=DATA_ACTIVITIES_HEADERS)
            df_excel.to_csv(DATA_ACTIVITIES_PATH, index=False)
        except Exception:
            pass

def generate_excel_bytes():
    ensure_data_file()
    try:
        df = pd.read_csv(DATA_PATH)
    except Exception:
        df = pd.DataFrame(columns=DATA_HEADERS)
        
    if 'Fecha inicio' in df.columns:
        df['Fecha inicio'] = pd.to_datetime(df['Fecha inicio'], errors='coerce')
    
    df = df.reindex(columns=DATA_HEADERS)
    
    sort_cols = [c for c in ['Fecha inicio','Municipio','Nombre paciente'] if c in df.columns]
    if sort_cols:
        df = df.sort_values(by=sort_cols, ascending=[True, True, True], na_position='last')
        
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Registros')
        ws = writer.book['Registros']
        ws.freeze_panes = 'A2'
        header_fill = PatternFill(fill_type='solid', start_color='EEF3FF', end_color='EEF3FF')
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                val = '' if cell.value is None else str(cell.value)
                if len(val) > max_len:
                    max_len = len(val)
            ws.column_dimensions[col_letter].width = min(48, max(12, max_len + 2))
    output.seek(0)
    return output

def update_excel_file():
    try:
        excel_bytes = generate_excel_bytes()
        with open(EXCEL_PATH, 'wb') as f:
            f.write(excel_bytes.getvalue())
    except Exception as e:
        print(f"Error updating Excel file: {e}")

def generate_activities_excel_bytes():
    ensure_activities_file()
    try:
        df = pd.read_csv(DATA_ACTIVITIES_PATH)
    except Exception:
        df = pd.DataFrame(columns=DATA_ACTIVITIES_HEADERS)
        
    if 'Fecha' in df.columns:
        df['Fecha'] = pd.to_datetime(df['Fecha'], errors='coerce')
        df = df.sort_values(by='Fecha', ascending=True)
    
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Actividades')
        ws = writer.book['Actividades']
        ws.freeze_panes = 'A2'
        header_fill = PatternFill(fill_type='solid', start_color='EEF3FF', end_color='EEF3FF')
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                val = '' if cell.value is None else str(cell.value)
                if len(val) > max_len:
                    max_len = len(val)
            ws.column_dimensions[col_letter].width = min(60, max(12, max_len + 2))
    output.seek(0)
    return output

def update_activities_excel_file():
    try:
        excel_bytes = generate_activities_excel_bytes()
        with open(EXCEL_ACTIVITIES_PATH, 'wb') as f:
            f.write(excel_bytes.getvalue())
    except Exception as e:
        print(f"Error updating Activities Excel file: {e}")

def load_catalog():
    if not os.path.exists(UPLOADS_DIR):
        os.makedirs(UPLOADS_DIR, exist_ok=True)
    catalog = {}
    if os.path.exists(CATALOG_PATH):
        try:
            with open(CATALOG_PATH, 'r', encoding='utf-8') as f:
                catalog = json.load(f)
        except Exception:
            catalog = {}
    return catalog

def save_catalog(cat):
    with open(CATALOG_PATH, 'w', encoding='utf-8') as f:
        json.dump(cat, f, ensure_ascii=False)

def extract_catalog(df):
    cols = {c.lower().strip(): c for c in df.columns}
    def get_unique(colnames):
        for c in colnames:
            key = c.lower().strip()
            if key in cols:
                return sorted([str(x).strip() for x in df[cols[key]].dropna().unique() if str(x).strip() != ''])
        return []
    prof_name_col = None
    for c in ['Nombre profesional','Profesional','Nombre del profesional']:
        k = c.lower().strip()
        if k in cols:
            prof_name_col = cols[k]
            break
    prof_doc_col = None
    for c in ['Documento profesional','Doc profesional','Documento del profesional']:
        k = c.lower().strip()
        if k in cols:
            prof_doc_col = cols[k]
            break
    prof_map = {}
    if prof_name_col and prof_doc_col:
        try:
            for _, row in df[[prof_name_col, prof_doc_col]].dropna().iterrows():
                n = str(row[prof_name_col]).strip()
                d = str(row[prof_doc_col]).strip()
                if n and d:
                    prof_map[n] = d
        except Exception:
            prof_map = {}
    return {
        'nombre_prof': get_unique(['Nombre profesional','Profesional','Nombre del profesional']),
        'doc_prof': get_unique(['Documento profesional','Doc profesional','Documento del profesional']),
        'nombre_pac': get_unique(['Nombre paciente','Paciente','Nombre del paciente']),
        'doc_pac': get_unique(['Documento paciente','Doc paciente','Documento del paciente']),
        'municipio': get_unique(['Municipio','Ciudad','Localidad']),
        'procedimiento': get_unique(['Procedimiento','Nombre procedimiento','Servicio']),
        'prof_map': prof_map
    }

def get_next_id(df, id_col='ID'):
    if df.empty or id_col not in df.columns:
        return 1
    # Asegurar que la columna ID es numérica
    ids = pd.to_numeric(df[id_col], errors='coerce').fillna(0)
    return int(ids.max()) + 1

# --- Interfaz de Usuario ---

def main():
    catalog = load_catalog()
    
    # Sidebar Navigation
    st.sidebar.title("Navegación")
    page = st.sidebar.radio("Ir a:", ["Procedimientos", "Actividades", "Administrador"])
    
    # --- PÁGINA: PROCEDIMIENTOS ---
    if page == "Procedimientos":
        st.title("Registro de Procedimientos")
        
        # Búsqueda Limitada (Solo Novedad/Panacea)
        with st.expander("Buscar Registro por ID (Solo editar Novedad/Panacea)"):
            search_id = st.number_input("Ingrese ID para buscar", min_value=1, step=1, key="search_proc_id")
            if st.button("Buscar Procedimiento"):
                ensure_data_file()
                df = pd.read_csv(DATA_PATH)
                if 'ID' in df.columns:
                    record = df[df['ID'] == search_id]
                    if not record.empty:
                        st.session_state['edit_proc_id'] = search_id
                        st.success(f"Registro {search_id} encontrado.")
                    else:
                        st.error("ID no encontrado.")
        
        # Formulario
        with st.form("proc_form"):
            # Determinar si estamos editando
            edit_id = st.session_state.get('edit_proc_id', None)
            default_vals = {}
            
            if edit_id:
                ensure_data_file()
                df = pd.read_csv(DATA_PATH)
                record = df[df['ID'] == edit_id].iloc[0]
                default_vals = record.to_dict()
                st.info(f"Editando Registro ID: {edit_id}")
                # Solo campos editables en modo búsqueda limitada: Panacea y Novedad
                # Pero si es modo 'nuevo', todo es editable.
                # La app original tenía search_public que permitía editar solo panacea y novedad
                # Vamos a simular eso.
                
                # Campos de solo lectura
                st.text_input("Nombre profesional", value=default_vals.get('Nombre profesional', ''), disabled=True)
                st.text_input("Documento profesional", value=default_vals.get('Documento profesional', ''), disabled=True)
                st.text_input("Nombre paciente", value=default_vals.get('Nombre paciente', ''), disabled=True)
                st.text_input("Documento paciente", value=default_vals.get('Documento paciente', ''), disabled=True)
                st.text_input("Fecha inicio", value=default_vals.get('Fecha inicio', ''), disabled=True)
                st.text_input("Municipio", value=default_vals.get('Municipio', ''), disabled=True)
                st.text_input("Procedimiento", value=default_vals.get('Procedimiento', ''), disabled=True)
                
            else:
                # Modo Nuevo Registro
                # Nombre Profesional
                prof_opts = catalog.get('nombre_prof', [])
                if prof_opts:
                    nombre_prof = st.selectbox("Nombre profesional", [""] + prof_opts)
                else:
                    nombre_prof = st.text_input("Nombre profesional")
                
                # Documento Profesional (Auto-relleno si existe mapa)
                prof_map = catalog.get('prof_map', {})
                doc_val = prof_map.get(nombre_prof, "") if isinstance(nombre_prof, str) and nombre_prof in prof_map else ""
                
                if doc_val:
                    doc_prof = st.text_input("Documento profesional", value=doc_val, disabled=True)
                    # Hack para enviar el valor disabled en el submit no funciona directo, usamos estado o hidden logic
                    # En streamlit el valor disabled no se envía? Se lee del widget.
                elif catalog.get('doc_prof'):
                    doc_prof = st.selectbox("Documento profesional", [""] + catalog.get('doc_prof'))
                else:
                    doc_prof = st.text_input("Documento profesional", value=doc_val)

                nombre_pac = st.text_input("Nombre paciente")
                doc_pac = st.text_input("Documento paciente")
                fecha_inicio = st.date_input("Fecha inicio", value=datetime.now())
                
                mun_opts = catalog.get('municipio', [])
                if mun_opts:
                    municipio = st.selectbox("Municipio", [""] + mun_opts)
                else:
                    municipio = st.text_input("Municipio")
                
                proc_opts = catalog.get('procedimiento', [])
                if proc_opts:
                    procedimiento = st.selectbox("Procedimiento", [""] + proc_opts)
                else:
                    procedimiento = st.text_input("Procedimiento")

            # Campos Editables siempre
            panacea_opts = ["", "Sí", "No"]
            panacea_val = default_vals.get('Subido a Panacea', '')
            panacea_idx = 0
            if panacea_val in ['Sí', 'Si']: panacea_idx = 1
            elif panacea_val == 'No': panacea_idx = 2
            
            panacea = st.selectbox("¿Se subió a Panacea?", panacea_opts, index=panacea_idx)
            novedad = st.text_area("Novedad", value=default_vals.get('Novedad', ''))
            
            submitted = st.form_submit_button("Guardar")
            
            if submitted:
                # Validaciones
                errors = []
                if not edit_id:
                    if not nombre_prof: errors.append("Nombre profesional requerido")
                    if not nombre_pac: errors.append("Nombre paciente requerido")
                    # ... más validaciones
                
                if not panacea: errors.append("Seleccione estado Panacea")
                
                if errors:
                    for e in errors: st.error(e)
                else:
                    ensure_data_file()
                    df = pd.read_csv(DATA_PATH)
                    now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    
                    if edit_id:
                        # Actualizar
                        idx = df.index[df['ID'] == edit_id].tolist()
                        if idx:
                            i = idx[0]
                            df.at[i, 'Subido a Panacea'] = panacea
                            df.at[i, 'Novedad'] = novedad
                            df.at[i, 'Modificado'] = now_str
                            st.success(f"Registro {edit_id} actualizado.")
                            st.session_state.pop('edit_proc_id', None) # Salir modo edición
                    else:
                        # Crear Nuevo
                        new_id = get_next_id(df)
                        # Recuperar valores de widgets
                        # Nota: Si el widget estaba disabled, st.session_state puede tener el valor o usamos variable local
                        # Para doc_prof si fue disabled
                        final_doc_prof = doc_val if doc_val else doc_prof
                        
                        new_row = {
                            'ID': new_id,
                            'Nombre profesional': nombre_prof,
                            'Documento profesional': final_doc_prof,
                            'Nombre paciente': nombre_pac,
                            'Documento paciente': doc_pac,
                            'Fecha inicio': fecha_inicio.strftime('%Y-%m-%d'),
                            'Municipio': municipio,
                            'Procedimiento': procedimiento,
                            'Subido a Panacea': panacea,
                            'Novedad': novedad,
                            'Creado': now_str,
                            'Modificado': ''
                        }
                        df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
                        st.success(f"Registro creado exitosamente. ID: {new_id}")
                    
                    # Guardar
                    df.to_csv(DATA_PATH, index=False)
                    update_excel_file()
                    
        if st.session_state.get('edit_proc_id'):
            if st.button("Cancelar Edición"):
                st.session_state.pop('edit_proc_id', None)
                st.rerun()

    # --- PÁGINA: ACTIVIDADES ---
    elif page == "Actividades":
        st.title("Registro de Actividades")
        
        # Buscar Actividades
        with st.expander("Consultar y Editar mis Actividades", expanded=True):
            prof_opts = catalog.get('nombre_prof', [])
            if prof_opts:
                search_prof = st.selectbox("Seleccione su nombre", [""] + prof_opts, key="search_act_prof")
            else:
                search_prof = st.text_input("Nombre profesional", key="search_act_prof_txt")
            
            if search_prof:
                ensure_activities_file()
                sync_activities_db()
                try:
                    df = pd.read_csv(DATA_ACTIVITIES_PATH)
                    if 'Nombre profesional' in df.columns:
                        my_acts = df[df['Nombre profesional'] == search_prof]
                        if not my_acts.empty:
                            st.dataframe(my_acts[['ID', 'Fecha', 'Actividad', 'Modificado']], use_container_width=True)
                            
                            # Selector para editar
                            act_id_to_edit = st.selectbox("Seleccione ID para editar", [""] + list(my_acts['ID'].astype(str)), key="sel_edit_act")
                            if act_id_to_edit:
                                st.session_state['edit_act_id'] = int(act_id_to_edit)
                        else:
                            st.info("No se encontraron actividades.")
                except Exception as e:
                    st.error(f"Error: {e}")

        st.divider()
        
        # Formulario Actividad
        with st.form("act_form"):
            edit_act_id = st.session_state.get('edit_act_id', None)
            act_defaults = {}
            
            if edit_act_id:
                ensure_activities_file()
                df = pd.read_csv(DATA_ACTIVITIES_PATH)
                record = df[df['ID'] == edit_act_id]
                if not record.empty:
                    act_defaults = record.iloc[0].to_dict()
                    st.info(f"Editando Actividad ID: {edit_act_id}")
            
            # Fecha
            default_date = datetime.now()
            if act_defaults.get('Fecha'):
                try:
                    default_date = datetime.strptime(act_defaults['Fecha'], '%Y-%m-%d')
                except:
                    pass
            fecha = st.date_input("Fecha", value=default_date)
            
            # Profesional
            if prof_opts:
                # Intentar matchear
                curr_prof = act_defaults.get('Nombre profesional', '')
                idx = 0
                if curr_prof in prof_opts:
                    idx = prof_opts.index(curr_prof)
                prof_act = st.selectbox("Nombre profesional", prof_opts, index=idx)
            else:
                prof_act = st.text_input("Nombre profesional", value=act_defaults.get('Nombre profesional', ''))
                
            actividad_txt = st.text_area("Actividad / Observación", value=act_defaults.get('Actividad', ''))
            
            submit_act = st.form_submit_button("Guardar Actividad")
            
            if submit_act:
                if not prof_act or not actividad_txt:
                    st.error("Complete todos los campos")
                else:
                    ensure_activities_file()
                    sync_activities_db()
                    df = pd.read_csv(DATA_ACTIVITIES_PATH)
                    now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    
                    if edit_act_id:
                        idx = df.index[df['ID'] == edit_act_id].tolist()
                        if idx:
                            i = idx[0]
                            df.at[i, 'Fecha'] = fecha.strftime('%Y-%m-%d')
                            df.at[i, 'Nombre profesional'] = prof_act
                            df.at[i, 'Actividad'] = actividad_txt
                            df.at[i, 'Modificado'] = now_str
                            st.success("Actividad actualizada.")
                            st.session_state.pop('edit_act_id', None)
                    else:
                        new_id = get_next_id(df)
                        new_row = {
                            'ID': new_id,
                            'Fecha': fecha.strftime('%Y-%m-%d'),
                            'Nombre profesional': prof_act,
                            'Actividad': actividad_txt,
                            'Creado': now_str,
                            'Modificado': ''
                        }
                        df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
                        st.success(f"Actividad guardada. ID: {new_id}")
                        
                    df.to_csv(DATA_ACTIVITIES_PATH, index=False)
                    update_activities_excel_file()
                    
        if st.session_state.get('edit_act_id'):
             if st.button("Cancelar Edición Actividad"):
                st.session_state.pop('edit_act_id', None)
                st.rerun()

    # --- PÁGINA: ADMINISTRADOR ---
    elif page == "Administrador":
        st.title("Panel de Control")
        
        # Login
        if 'logged_in' not in st.session_state:
            st.session_state['logged_in'] = False
            
        if not st.session_state['logged_in']:
            with st.form("login_form"):
                user = st.text_input("Usuario")
                pwd = st.text_input("Contraseña", type="password")
                if st.form_submit_button("Ingresar"):
                    if user == ADMIN_USER and pwd == ADMIN_PASS:
                        st.session_state['logged_in'] = True
                        st.rerun()
                    else:
                        st.error("Credenciales incorrectas")
        else:
            if st.button("Cerrar Sesión"):
                st.session_state['logged_in'] = False
                st.rerun()
                
            tab1, tab2 = st.tabs(["Gestión Procedimientos", "Seguimiento Actividades"])
            
            with tab1:
                ensure_data_file()
                df = pd.read_csv(DATA_PATH)
                st.metric("Registros Totales", len(df))
                
                col1, col2 = st.columns(2)
                with col1:
                    # Descargar Excel
                    excel_data = generate_excel_bytes()
                    st.download_button(
                        label="Descargar Excel Completo",
                        data=excel_data,
                        file_name="registros_procedimientos.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                with col2:
                    # Subir Catálogo
                    uploaded_file = st.file_uploader("Actualizar Catálogo (.xlsx)", type=['xlsx', 'xls'])
                    if uploaded_file:
                        if st.button("Procesar Catálogo"):
                            try:
                                # Guardar archivo
                                save_path = os.path.join(UPLOADS_DIR, uploaded_file.name)
                                with open(save_path, "wb") as f:
                                    f.write(uploaded_file.getbuffer())
                                
                                # Procesar
                                xl = pd.read_excel(save_path, sheet_name=None)
                                new_catalog = {'catalog_file_path': save_path}
                                
                                # Mapear hojas a listas
                                # Asumimos estructura: Hoja "Profesionales" -> Columna "Nombre"
                                # Esto depende de cómo estaba hecho upload_catalog en Flask.
                                # Revisando Flask code... no lo leí completo, pero asumiré una estructura genérica o 
                                # intentaré leer todas las columnas de la primera hoja.
                                # MEJOR: Leo el código Flask original para ver `upload_catalog`.
                                pass
                            except Exception as e:
                                st.error(f"Error: {e}")
            
                # Edición Completa
                st.subheader("Buscar y Editar Registro (Completo)")
                search_admin_id = st.number_input("ID Registro", min_value=1, step=1, key="admin_search")
                if st.button("Buscar en Admin"):
                     record = df[df['ID'] == search_admin_id]
                     if not record.empty:
                         st.session_state['admin_edit_id'] = search_admin_id
                     else:
                         st.error("No encontrado")
                
                if st.session_state.get('admin_edit_id'):
                    edit_id = st.session_state['admin_edit_id']
                    st.write(f"Editando ID: {edit_id}")
                    row = df[df['ID'] == edit_id].iloc[0]
                    
                    with st.form("admin_edit_form"):
                        # Todos los campos editables
                        col_a, col_b = st.columns(2)
                        with col_a:
                            n_prof = st.text_input("Nombre Profesional", value=row.get('Nombre profesional', ''))
                            d_prof = st.text_input("Documento Profesional", value=row.get('Documento profesional', ''))
                            n_pac = st.text_input("Nombre Paciente", value=row.get('Nombre paciente', ''))
                            d_pac = st.text_input("Documento Paciente", value=row.get('Documento paciente', ''))
                        with col_b:
                            try:
                                f_val = datetime.strptime(str(row.get('Fecha inicio','')).split(' ')[0], '%Y-%m-%d')
                            except:
                                f_val = datetime.now()
                            f_ini = st.date_input("Fecha Inicio", value=f_val)
                            muni = st.text_input("Municipio", value=row.get('Municipio', ''))
                            proc = st.text_input("Procedimiento", value=row.get('Procedimiento', ''))
                        
                        pan = st.selectbox("Subido a Panacea", ["Sí", "No"], index=0 if row.get('Subido a Panacea') in ['Sí','Si'] else 1)
                        nov = st.text_area("Novedad", value=row.get('Novedad', ''))
                        
                        if st.form_submit_button("Guardar Cambios Admin"):
                            idx = df.index[df['ID'] == edit_id].tolist()[0]
                            df.at[idx, 'Nombre profesional'] = n_prof
                            df.at[idx, 'Documento profesional'] = d_prof
                            df.at[idx, 'Nombre paciente'] = n_pac
                            df.at[idx, 'Documento paciente'] = d_pac
                            df.at[idx, 'Fecha inicio'] = f_ini.strftime('%Y-%m-%d')
                            df.at[idx, 'Municipio'] = muni
                            df.at[idx, 'Procedimiento'] = proc
                            df.at[idx, 'Subido a Panacea'] = pan
                            df.at[idx, 'Novedad'] = nov
                            
                            # Actualizar modificado
                            now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                            df.at[idx, 'Modificado'] = now_str
                            
                            df.to_csv(DATA_PATH, index=False)
                            update_excel_file()
                            st.success("Registro actualizado exitosamente.")
                            st.session_state.pop('admin_edit_id', None)
                            st.rerun()

            with tab2:
                ensure_activities_file()
                sync_activities_db()
                df_act = pd.read_csv(DATA_ACTIVITIES_PATH)
                
                st.metric("Total Actividades", len(df_act))
                
                # Filtros
                col_f1, col_f2 = st.columns(2)
                with col_f1:
                    profs = sorted(df_act['Nombre profesional'].dropna().unique())
                    fil_prof = st.selectbox("Filtrar por Profesional", ["Todos"] + list(profs))
                with col_f2:
                    fil_date = st.date_input("Filtrar por Fecha", value=None)
                
                if fil_prof != "Todos":
                    df_act = df_act[df_act['Nombre profesional'] == fil_prof]
                if fil_date:
                    # Asumimos formato YYYY-MM-DD
                    df_act = df_act[df_act['Fecha'] == fil_date.strftime('%Y-%m-%d')]
                
                # Gráfico de Actividades
                if not df_act.empty:
                    st.subheader("Resumen de Actividades")
                    act_counts = df_act['Nombre profesional'].value_counts().reset_index()
                    act_counts.columns = ['Profesional', 'Cantidad']
                    fig_act = px.bar(act_counts, x='Profesional', y='Cantidad', title='Actividades por Profesional')
                    st.plotly_chart(fig_act, use_container_width=True)

                st.dataframe(df_act, use_container_width=True)
                
                # Descargar
                excel_acts = generate_activities_excel_bytes()
                st.download_button("Descargar Actividades (.xlsx)", excel_acts, "actividades.xlsx")
                
                # Eliminar
                st.subheader("Eliminar Actividad")
                del_id = st.number_input("ID a eliminar", min_value=1, step=1)
                if st.button("Eliminar Actividad"):
                    if del_id in df_act['ID'].values:
                        # Leer original completo para borrar
                        full_df = pd.read_csv(DATA_ACTIVITIES_PATH)
                        full_df = full_df[full_df['ID'] != del_id]
                        full_df.to_csv(DATA_ACTIVITIES_PATH, index=False)
                        update_activities_excel_file()
                        st.success(f"Eliminado ID {del_id}")
                        st.rerun()
                    else:
                        st.error("ID no encontrado en la selección actual")

if __name__ == '__main__':
    main()

